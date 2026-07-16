"""
Gera uma planilha Excel com todas as SÉRIES TEMPORAIS (caminhos raiz->folha) da
árvore de linhagem contida no CSV de tracking.

Cada série temporal = um caminho da célula raiz até uma ponta (folha) da árvore.
Cada LINHA do bloco = uma célula (track_id) desse caminho. Blocos são separados
por uma linha em branco e o cabeçalho é repetido a cada série.

Colunas por célula:
  Frames:        Initial / End                    -> min/max de 'frame'
  Intermitotic:  Time                             -> 'lifetime'
  Nuclear Area:  Average/SD/Inclination/R2        -> coluna 'area_px'
  Nuclear NII:   Average/SD/Inclination/R2        -> coluna 'nii'
  ERK-KTR:       1st4/last4/Average/SD/Incl/R2    -> coluna 'green_preprocessed_cn_ratio'

Inclination = coeficiente angular da regressão linear (valor ~ frame);
R2 = coeficiente de determinação da mesma regressão.
"""
import argparse
from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AREA_COL = "area_px"
NII_COL = "nii"
ERK_COL = "green_preprocessed_cn_ratio"


def linreg(x, y):
    """Retorna (inclinação, R2) da reta y ~ x. NaN se < 2 pontos válidos."""
    mask = np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    if len(x) < 2 or np.ptp(x) == 0:
        return np.nan, np.nan
    slope, intercept = np.polyfit(x, y, 1)
    y_pred = slope * x + intercept
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r2 = 1 - ss_res / ss_tot if ss_tot != 0 else np.nan
    return slope, r2


def cell_metrics(g):
    """Métricas de um track_id (sub-DataFrame ordenado por frame)."""
    frame = g["frame"].to_numpy(float)
    area = g[AREA_COL].to_numpy(float)
    nii = g[NII_COL].to_numpy(float)
    erk = g[ERK_COL].to_numpy(float)
    erk_valid = erk[np.isfinite(erk)]

    a_slope, a_r2 = linreg(frame, area)
    n_slope, n_r2 = linreg(frame, nii)
    e_slope, e_r2 = linreg(frame, erk)

    return [
        int(g["frame"].min()), int(g["frame"].max()),   # Frames Initial/End
        g["lifetime"].dropna().iloc[0] if g["lifetime"].notna().any() else np.nan,  # Intermitotic
        np.nanmean(area), np.nanstd(area), a_slope, a_r2,               # Nuclear Area
        np.nanmean(nii), np.nanstd(nii), n_slope, n_r2,                 # Nuclear NII
        np.nanmean(erk_valid[:4]) if len(erk_valid) else np.nan,        # ERK 1st 4
        np.nanmean(erk_valid[-4:]) if len(erk_valid) else np.nan,       # ERK last 4
        np.nanmean(erk), np.nanstd(erk), e_slope, e_r2,                 # ERK avg/SD/reg
    ]


def enumerate_paths(root, children):
    """Todos os caminhos raiz->folha como listas de track_id (DFS iterativo)."""
    paths, stack = [], [[root]]
    while stack:
        path = stack.pop()
        kids = children.get(path[-1], [])
        if not kids:
            paths.append(path)
        else:
            for k in kids:
                stack.append(path + [k])
    return paths


# --- Escrita da planilha ---
GROUPS = [
    ("Frames", ["Initial", "End"], "FFFFFF"),
    ("Intermitotic", ["Time"], "FFFFFF"),
    ("Nuclear Area (NA)", ["Average", "SD", "Inclination", "R2"], "E8D3E8"),
    ("Nuclear NII", ["Average", "SD", "Inclination", "R2"], "CFE7F5"),
    ("ERK-KTR", ["1st 4 frames", "last 4 frames", "Average", "SD", "Inclination", "R2"], "D3EAD3"),
]
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_headers(ws, row):
    """Escreve as 2 linhas de cabeçalho a partir de `row`. Retorna próxima linha livre."""
    ws.cell(row=row, column=1, value="Séries temporais").font = Font(bold=True)
    col = 2
    for title, subs, color in GROUPS:
        fill = PatternFill("solid", fgColor=color)
        if len(subs) > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + len(subs) - 1)
        top = ws.cell(row=row, column=col, value=title)
        top.font = Font(bold=True)
        top.alignment = Alignment(horizontal="center")
        top.fill = fill
        for i, sub in enumerate(subs):
            c = ws.cell(row=row + 1, column=col + i, value=sub)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            c.fill = fill
            c.border = BORDER
        col += len(subs)
    return row + 2


def tree_layout(root, children):
    """Posições (x, y) dos nós: folhas espalhadas em x, pai = média dos filhos; y = -profundidade."""
    pos, leaf_x = {}, [0]

    def place(node, depth):
        kids = children.get(node, [])
        if not kids:
            x = leaf_x[0]
            leaf_x[0] += 1
        else:
            x = float(np.mean([place(k, depth + 1) for k in kids]))
        pos[node] = (x, -depth)
        return x

    place(root, 0)
    return pos


def plot_tree(root, children, outcome_of):
    """Desenha a árvore de linhagem e retorna um PNG em memória (BytesIO)."""
    pos = tree_layout(root, children)
    n_leaves = sum(1 for n in pos if not children.get(n))
    depth = max(-y for _, y in pos.values()) + 1
    fig, ax = plt.subplots(figsize=(max(6, n_leaves * 0.7), max(4, depth * 1.0)))

    for node, (x, y) in pos.items():
        for k in children.get(node, []):
            kx, ky = pos[k]
            ax.plot([x, kx], [y, ky], "-", color="0.6", lw=1, zorder=1)
    for node, (x, y) in pos.items():
        color = "#d98c8c" if outcome_of.get(node) == "Mitosis" else "#8cb6d9"
        ax.scatter([x], [y], s=520, color=color, edgecolors="black", zorder=2)
        ax.text(x, y, str(node), ha="center", va="center", fontsize=7, zorder=3)

    ax.set_title(f"Árvore de linhagem — raiz {root}")
    ax.axis("off")
    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def build_workbook(df, root):
    children = {p: sorted(g.unique()) for p, g in df.groupby("parent_id")["track_id"]}
    groups = {tid: g.sort_values("frame") for tid, g in df.groupby("track_id")}

    if root not in groups:
        raise ValueError(f"track_id raiz {root} não existe no CSV.")

    paths = enumerate_paths(root, children)
    paths.sort()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Series_raiz_{root}"

    row = 1
    for path in paths:
        row = write_headers(ws, row)
        for tid in path:
            vals = cell_metrics(groups[tid])
            ws.cell(row=row, column=1, value=tid)
            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=2 + i, value=None if pd.isna(v) else v)
                if i >= 2:  # colunas numéricas de features
                    cell.number_format = "0.000"
            row += 1
        row += 1  # linha em branco entre séries

    for i, w in enumerate([16, 8, 8, 8, 10, 8, 10, 8, 10, 8, 10, 8, 12, 12, 10, 8, 10, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    outcome_of = df.groupby("track_id")["outcome"].last().to_dict()
    ws_tree = wb.create_sheet(f"Arvore_raiz_{root}")
    ws_tree.add_image(XLImage(plot_tree(root, children, outcome_of)), "A1")

    return wb, len(paths)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input_csv", help="CSV de tracking com features (tracks_com_rois_...).")
    ap.add_argument("output_xlsx", help="Caminho da planilha Excel de saída.")
    ap.add_argument("--root", type=int, default=1, help="track_id raiz (padrão: 1).")
    args = ap.parse_args()

    df = pd.read_csv(args.input_csv)
    wb, n = build_workbook(df, args.root)
    wb.save(args.output_xlsx)
    print(f"{n} séries temporais escritas em {args.output_xlsx} (raiz={args.root}).")


if __name__ == "__main__":
    main()
